'''
Created on Jul 12, 2020

@author: stand
'''
import time

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import pandas as pd
import re
import datetime
import sys

import os
import shutil
from os import listdir
from os.path import isfile, join
from openpyxl import load_workbook


def Activate_driver(pathin, driverpath= r'D:\Downloads\edgedriver_win64\chromedriver.exe'):
    chromeOptions = webdriver.ChromeOptions()
    prefs = {"download.default_directory" : pathin}
    chromeOptions.add_experimental_option("prefs",prefs)
    chromeOptions.add_argument("--headless")

    driver = webdriver.Chrome(options=chromeOptions, executable_path = driverpath)
    return driver

def Download_ticker(driver, ticker):
    start_url = "https://finance.yahoo.com/quote/"  + str(ticker) + "/history?p=" + str(ticker)
    driver.get(start_url)


    # time.sleep(5)
    # elements = driver.find_elements_by_xpath("//a[contains(@href, 'https://query1.finance.yahoo.com/v7/finance/download/')]")
    try:
        time.sleep(1)
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//a[contains(@href, 'https://query1.finance.yahoo.com/v7/finance/download/')]"))
        )

        elements = driver.find_elements_by_xpath("//a[contains(@href, 'https://query1.finance.yahoo.com/v7/finance/download/')]")
        elements[0].click()
        time.sleep(5)
    except:
        pass

def Default_sort_function(df, column='Date', ascending=False):
    return df.sort_values(by=[column], inplace=True, ascending=ascending)


def Process_set(excel_base_path, file, path_scan, copy_start_row=2, copy_start_column=26):
    m = re.match('(.*)\.(.*?)$', str(file))
    if m:
        name = m.group(1).upper()
        extension = m.group(2).lower()

    df = None
    if extension == 'csv':
        df = pd.read_csv(join(path_scan,file))
    if extension.endswith('xls'):
        df = pd.read_excel(join(path_scan,file))
    
    if not df.empty:
        Default_sort_function(df)
        
        wb = load_workbook(filename = excel_base_path)
        ws = wb.get_sheet_by_name(name)

        col_index = copy_start_column
        for col,data in df.items():
            row_index = copy_start_row
            for val in data.values:
                if col == 'Date':
                    date_value = datetime.datetime.strptime(str(val), "%Y-%m-%d").strftime('%m/%d/%Y')
                    date_value = datetime.datetime.strptime(str(date_value), '%m/%d/%Y')
#                     date_value = datetime.datetime.strptime(str(val), "%Y-%m-%d").date()
#                     date_value = '{0}/{1}/{2}'.format(date_value.month, date_value.day, date_value.year)
                    ws.cell(row=row_index,column=col_index).value = date_value
                else:
                    ws.cell(row=row_index,column=col_index).value = val
                row_index += 1
            col_index += 1
        
        wb.save(excel_base_path)    
            
def Scan_directory_and_process(excel_base_path, path_scan):
    files = [f for f in listdir(path_scan) if isfile(join(path_scan, f))]

    for file in files:
        Process_set(excel_base_path, file, path_scan)    
    
    

if __name__ == '__main__':
    print(sys.argv)
    driverpath = str(sys.argv[1])
    pathin = str(sys.argv[2])
    excel_path = str(sys.argv[3])
    tickers = sys.argv[4:]
    
    
    shutil.rmtree(pathin)
    os.mkdir(pathin)
    
    driver = Activate_driver(pathin, driverpath)
    for ticker in tickers:
        Download_ticker(driver, ticker)
        
    driver.quit()
    
    Scan_directory_and_process(excel_path, pathin)